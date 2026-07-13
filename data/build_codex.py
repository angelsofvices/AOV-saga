#!/usr/bin/env python3
"""
build_codex.py — parse the Game Codex xlsx into data/codex.json.

Reads:
  data/aov_game_codex_v11.1.xlsx
    ├─ Sheet "9 MASTER INDEX"     — 525+ species entries (canonical stats/types/moves)
    └─ Sheet "14 GAME EVOLUTION"  — evolution chains (Chain / Stage 1-4 / Evolve Lvs / Invented / Notes)

Writes:
  data/codex.json — every entry keyed by lowercase-name slug, with:
    id, name, role ('zyrex' | 'zyraxian'), class, tier, tierRoman, tierClass,
    type1, type2, archetype, base { hp/atk/def/spd/spc }, moves, worldsense,
    flavor, plus (when chained):
      chain (str)       — chain id
      chainIdx (int)    — 0-based stage index
      evolveTo (str?)   — next stage id (if not terminal)
      evolveLv (int?)   — level at which evolution triggers
      invented (bool)   — True if this stage was invented by us (not in codex canon)

Invented species get a scaled-down stat block derived from the next canonical
stage in the chain (so an invented Tier I hatchling below a Tier V Apex gets
a Tier I stat block extrapolated backward).

Run:
  python3 data/build_codex.py
"""
from __future__ import annotations
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / 'data' / 'aov_game_codex_v11.1.xlsx'
OUT  = ROOT / 'data' / 'codex.json'

ROMAN_TIERS = {'I':1,'II':2,'III':3,'IV':4,'V':5,'VI':6,'VII':7,'VIII':8,'IX':9,'X':10}
ZYREX_CLASSES = {'Beast', 'Beastmaster', 'Creature', 'Sovereign', 'Supreme', 'Antagonist', 'TBD', '—'}

# V2.39a — canonical renames applied to every parsed cell as a safety net.
CANON_RENAMES = [
    ("Ael'Tharion", 'Xenoxil'),
    ("AEL'THARION", 'XENOXIL'),
    ("Ael Tharion", 'Xenoxil'),
]
def apply_renames(v):
    if v is None: return v
    s = str(v)
    for old, new in CANON_RENAMES:
        s = s.replace(old, new)
    return s

TIER_CLASS_NAME = {
    1: 'Basic', 2: 'Core', 3: 'Champion', 4: 'Legend', 5: 'Apex',
    6: 'Pseudoimmortal', 7: 'Pseudoimmortal', 8: 'Immortal',
    9: 'Demigod', 10: 'God',
}

def slug(s):
    if not s: return None
    return re.sub(r'[^a-z0-9]+', '', str(s).lower())

def parse_move(cell):
    if not cell: return None
    s = str(cell).strip()
    m = re.match(r'^\s*([^—–\-]+?)\s*[—–\-]\s*(.*)$', s, re.DOTALL)
    name = m.group(1).strip() if m else s.split('\n')[0].strip()
    return {'name': name, 'raw': s}

def parse_master_index(ws):
    """Yield row-dicts for every species in the master index."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[2]
        if not name:
            continue
        tier_num = ROMAN_TIERS.get(str(row[3]).strip(), 1)
        cls = (row[1] or '').strip()
        role = 'zyrex' if cls in ZYREX_CLASSES else 'zyraxian'
        key = slug(name)
        if not key:
            continue
        # V2.27 — sheet 9 gained a Type 3 column at index 7 (0-based). All subsequent
        # columns shift right by 1.
        yield {
            'id': key,
            'name': str(name).strip(),
            'role': role,
            'class': cls,
            'tier': tier_num,
            'tierRoman': str(row[3]).strip() if row[3] else 'I',
            'tierClass': (row[4] or '').strip(),
            'type1': (row[5] or 'Beast').strip(),
            'type2': (row[6] or 'Beast').strip(),
            'type3': ((row[7] or '').strip() or None) if tier_num >= 6 else None,   # V2.27 — Tier 6+ only
            'archetype': (row[8] or '').strip(),
            'base': {
                'hp':  int(row[9])  if row[9]  else 66,
                'atk': int(row[10]) if row[10] else 66,
                'def': int(row[11]) if row[11] else 66,
                'spd': int(row[12]) if row[12] else 66,
                'spc': int(row[13]) if row[13] else 69,
            },
            'moves': {
                'a1': parse_move(row[15]),  'a1cast': (row[16] or '').strip(),
                'a2': parse_move(row[17]),  'a2cast': (row[18] or '').strip(),
                'a3': parse_move(row[19]),  'a3cast': (row[20] or '').strip(),
                'r':  {'name': (row[21] or '').strip()} if row[21] else None,
                'rcast': (row[22] or '').strip(),
                'invented': False,
            },
            'worldsense': apply_renames((row[23] or '').strip())[:400],
            'flavor':     apply_renames((row[24] or '').strip())[:400],
            'invented': False,
        }

def parse_evolution(ws):
    """Yield {chain, stages: [ids], evolveLvs: [ints], invented: set(codes)}.

    V2.27 note: the evolution section lives at the TOP of the sheet. Sheet 14
    was renamed to '14 GAME ROSTER' in V2.25 and now hosts a ROSTER section
    lower down. Stop parsing evolutions as soon as we hit the ROSTER divider
    or the header row 'Chain'."""
    for row in ws.iter_rows(min_row=4, values_only=True):
        chain = row[0]
        s1 = row[1]
        # Hard stop when we hit the ROSTER divider or its column headers.
        if chain and isinstance(chain, str) and ('GAME ROSTER' in chain or chain.strip() == '#'):
            break
        # Skip empty / notes / instruction rows.
        if not chain or not s1:
            continue
        # Only accept chain rows whose Stage 1 value is a proper species slug
        # (letters + digits, no spaces). The ROSTER rows have numeric IDs
        # (#) in col A and species IDs in col B — this filter also protects
        # against them if the divider check ever misses.
        if not isinstance(s1, str) or ' ' in s1 or not s1.strip():
            continue
        stages = []
        for i in range(1, 5):   # cols 1..4 = Stage 1..4
            v = row[i]
            if v is None or str(v).strip() == '':
                continue
            stages.append(slug(v))
        evolve_lvs_raw = row[5] or ''
        try:
            evolve_lvs = [int(x.strip()) for x in str(evolve_lvs_raw).split(',') if x.strip()]
        except ValueError:
            evolve_lvs = []
        invented_raw = row[6] or ''
        invented = {c.strip().upper() for c in str(invented_raw).split(',') if c.strip()}
        notes = row[7] or ''
        yield {
            'chain':     slug(chain),
            'stages':    stages,
            'evolveLvs': evolve_lvs,
            'invented':  invented,   # {'S1', 'S3', ...}
            'notes':     str(notes).strip(),
        }

def scale_base_down(canonical_base, from_tier, to_tier):
    """Scale a canonical stat block from its tier down toward a lower tier.

    V2.59 — Enforces the canonical T*333 rule: the resulting stat pool
    must equal to_tier * 333, preserving the relative weighting of the
    canonical block (so a magic-heavy Tier V species stays magic-heavy
    at Tier II).
    """
    if from_tier <= to_tier:
        return dict(canonical_base)
    target_total = to_tier * 333
    current_total = sum(canonical_base.get(k, 0) for k in ('hp','atk','def','spd','spc')) or 1
    scale = target_total / current_total
    scaled = {k: int(canonical_base.get(k, 0) * scale) for k in ('hp','atk','def','spd','spc')}
    # Trim floor drift onto SPC so the sum lands exactly at T*333.
    drift = target_total - sum(scaled.values())
    scaled['spc'] = max(1, scaled.get('spc', 0) + drift)
    return scaled

def synthesize_invented(id, chain, chain_idx, next_stage_entry, tier_guess):
    """Build a plausible codex entry for a stage that has no canonical row."""
    canonical_base = next_stage_entry['base'] if next_stage_entry else {'hp':100,'atk':100,'def':100,'spd':100,'spc':100}
    from_tier = next_stage_entry['tier'] if next_stage_entry else 5
    base = scale_base_down(canonical_base, from_tier, tier_guess)
    type1 = next_stage_entry['type1'] if next_stage_entry else 'Beast'
    type2 = next_stage_entry['type2'] if next_stage_entry else 'Beast'
    type3 = None   # V2.27 — invented lower stages never carry a 3rd type (Tier 6+ only)
    return {
        'id': id,
        'name': id.upper(),
        'role': 'zyrex',
        'class': next_stage_entry['class'] if next_stage_entry else 'Beast',
        'tier': tier_guess,
        'tierRoman': list(ROMAN_TIERS.keys())[tier_guess - 1],
        'tierClass': TIER_CLASS_NAME.get(tier_guess, 'Basic'),
        'type1': type1, 'type2': type2, 'type3': type3,
        'archetype': (next_stage_entry.get('archetype', '') if next_stage_entry else '') + ' (invented)',
        'base': base,
        'moves': {
            'a1': {'name': f'{id.title()} Nip', 'raw': 'invented — placeholder'},
            'a1cast': '',
            'a2': None, 'a2cast': '',
            'a3': None, 'a3cast': '',
            'r':  None, 'rcast': '',
            'invented': True,
        },
        'worldsense': f'[invented Stage {chain_idx+1} of chain "{chain}"]',
        'flavor':     f'{id.upper()} — invented lower form of chain "{chain}". Awaits canonical entry.',
        'invented': True,
    }

def guess_invented_tier(chain_stages, chain_idx, entries):
    """Given a chain's canonical stages, guess a reasonable tier for an
    invented stage at index chain_idx.

    Rule: Stage 1 (chain_idx == 0) always defaults to Tier I (hatchling).
    Middle-stage inventions scale by (offset × 2) below the next canonical.
    Above-canonical inventions add (offset × 2) up to Tier X.
    """
    if chain_idx == 0:
        return 1
    for j in range(chain_idx + 1, len(chain_stages)):
        e = entries.get(chain_stages[j])
        if e:
            step = (j - chain_idx) * 2
            return max(1, e['tier'] - step)
    for j in range(chain_idx - 1, -1, -1):
        e = entries.get(chain_stages[j])
        if e:
            step = (chain_idx - j) * 2
            return min(10, e['tier'] + step)
    return 1

def main():
    if not XLSX.exists():
        print(f'ERROR: {XLSX} missing', file=sys.stderr)
        sys.exit(1)
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=False)
    if '9 MASTER INDEX' not in wb.sheetnames:
        print('ERROR: sheet "9 MASTER INDEX" missing', file=sys.stderr)
        sys.exit(2)

    # 1. Parse all canonical entries.
    entries = {}
    for e in parse_master_index(wb['9 MASTER INDEX']):
        entries[e['id']] = e

    # 2. Parse evolution chains and stamp evolveTo/evolveLv/chain fields.
    # V2.25 renamed the sheet from "14 GAME EVOLUTION" to "14 GAME ROSTER".
    chains = []
    chain_sheet = None
    for candidate in ('14 GAME ROSTER', '14 GAME EVOLUTION'):
        if candidate in wb.sheetnames:
            chain_sheet = wb[candidate]; break
    if chain_sheet is not None:
        for c in parse_evolution(chain_sheet):
            chains.append(c)
            stages = c['stages']
            # Synthesize invented stages first so cross-refs work.
            for idx, sid in enumerate(stages):
                if sid not in entries:
                    tier_guess = guess_invented_tier(stages, idx, entries)
                    next_entry = None
                    for j in range(idx + 1, len(stages)):
                        if stages[j] in entries:
                            next_entry = entries[stages[j]]; break
                    entries[sid] = synthesize_invented(sid, c['chain'], idx, next_entry, tier_guess)
            # Stamp chain metadata.
            for idx, sid in enumerate(stages):
                e = entries[sid]
                e['chain'] = c['chain']
                e['chainIdx'] = idx
                # Default evolveLv rule: currentStage.tier × 10.
                if idx < len(stages) - 1:
                    next_id = stages[idx + 1]
                    if idx < len(c['evolveLvs']):
                        e['evolveLv'] = c['evolveLvs'][idx]
                    else:
                        e['evolveLv'] = e['tier'] * 10
                    e['evolveTo'] = next_id
                else:
                    e['evolveTo'] = None

    counts = {'total': len(entries), 'zyrex': 0, 'zyraxian': 0, 'invented': 0, 'chained': 0}
    for e in entries.values():
        counts[e['role']] = counts.get(e['role'], 0) + 1
        if e.get('invented'): counts['invented'] += 1
        if e.get('chain'):    counts['chained'] += 1

    out = {
        'source': f'Game Codex — parsed from {XLSX.name}',
        'generated_at_note': 'Regenerate via `python3 data/build_codex.py` after editing the xlsx.',
        'counts': counts,
        'chains': [{'chain': c['chain'], 'stages': c['stages'], 'evolveLvs': c['evolveLvs']} for c in chains],
        'entries': entries,
    }
    OUT.write_text(json.dumps(out, indent=1, ensure_ascii=False), encoding='utf-8')
    print(f'Wrote {OUT}  —  {counts}')
    for c in chains:
        print(f'  chain "{c["chain"]}": ' + ' → '.join(c['stages']))

    # V2.30 — Also emit codex.js, a script-tag-friendly slim payload the game
    # can hydrate SPECIES from at boot. Zyrex only (role == 'zyrex'); Zyraxian
    # humanoid classes are skipped (they'll drive NPC design later, not the
    # catchable roster).
    slim = {}
    NORM_TYPE = {'Void': 'Unknown-Void', 'Humanoid': 'Humanoid-Noid'}
    def norm_type(t):
        t = (t or '').strip()
        return NORM_TYPE.get(t, t) if t else None
    for id, e in entries.items():
        if e.get('role') != 'zyrex':
            continue
        base = e.get('base') or {}
        if not (base.get('hp') and base.get('atk')):
            continue
        slim[id] = {
            'id': id,
            'name': e['name'],
            'type': norm_type(e.get('type1')) or 'Beast',
            'type2': norm_type(e.get('type2')),
            'type3': norm_type(e.get('type3')) if e.get('tier', 1) >= 6 else None,
            'tier': e.get('tier') or 1,
            'base': {
                'hp':  base.get('hp')  or 66,
                'atk': base.get('atk') or 66,
                'def': base.get('def') or 66,
                'spd': base.get('spd') or 66,
                'spc': base.get('spc') or 69,
            },
            'archetype': e.get('archetype') or '',
            'chain':    e.get('chain'),
            'chainIdx': e.get('chainIdx'),
            'evolveTo': e.get('evolveTo'),
            'evolveLv': e.get('evolveLv'),
        }
    JS_OUT = ROOT / 'data' / 'codex.js'
    js = 'window.CODEX_ZYREX = ' + json.dumps(slim, ensure_ascii=False) + ';'
    JS_OUT.write_text(js, encoding='utf-8')
    print(f'Wrote {JS_OUT}  —  {len(slim)} Zyrex entries')

    # V2.59 — Canon audit: every species's base stat pool must equal
    # tier * 333.  Warn on any regression so future codex edits don't
    # silently drift out of canon.
    print()
    print('T*333 stat-pool audit:')
    mismatches = []
    for id_, e in slim.items():
        tier = e.get('tier') or 1
        b = e.get('base') or {}
        total = sum(b.get(k, 0) for k in ('hp','atk','def','spd','spc'))
        target = tier * 333
        if total != target:
            mismatches.append((id_, tier, total, target))
    if not mismatches:
        print(f'  OK — all {len(slim)} Zyrex satisfy pool = tier * 333.')
    else:
        print(f'  {len(mismatches)} mismatch(es):')
        for id_, tier, total, target in mismatches:
            print(f'    {id_:20s} T{tier}  pool={total:5d}  (needs {target}, delta {target-total:+d})')

if __name__ == '__main__':
    main()
